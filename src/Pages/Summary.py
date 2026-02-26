import os
import pandas as pd
from tkinter import *
from openpyxl import *
from tkinter import messagebox
from tkinter.ttk import Separator

from Utils.readSettings import readSettings
from Utils.prass import PRASS
from Utils.emailAlert import sendErrorAlertEmail

class Summary(readSettings):
    def __init__(self,root,inData,filePath,Wscreen,Hscreen):
        super().__init__()
        self.initialize(inData,filePath,Wscreen,Hscreen)
        self.reset(root)

    def reset(self,root):
        self.root = Toplevel(root)
        self.win_config()
        self.widgets()
        self.load_SumData(root)
        self.root.grab_set()
        self.root.mainloop()

    def initialize(self,inData,filePath,Wscreen,Hscreen):
        self.res = True
        self.filePath = filePath
        self.inData = inData
        self.Wscreen,self.Hscreen = Wscreen,Hscreen

    def win_config(self):
        self.root.title("Summary Window")
        self.frame = Frame(self.root)
        self.frame.pack()

    def widgets(self):
        # Labelframe for Lot Number Data and Block Weight
        ####################################################################################################

        lotNumCont = LabelFrame(self.frame, bd=5, relief=FLAT)
        lotNumCont.grid(row=0, column=0, padx=3, pady=1,sticky=W)

        lotNumTxt = Label(lotNumCont, text="Lot Number: ", font=self.font['S'])
        lotNumTxt.grid(row=0, column=0, padx=3, pady=1, sticky=W)

        lotNum = Label(lotNumCont, text=self.inData[0], font=self.font['S'])
        lotNum.grid(row=0, column=1, padx=3, pady=1)
        
        # Add Block Weight display
        blkWeightTxt = Label(lotNumCont, text="Block Weight: ", font=self.font['S'])
        blkWeightTxt.grid(row=0, column=2, padx=(20,3), pady=1, sticky=W)
        
        blkWeight = Label(lotNumCont, text=self.inData[4], font=self.font['S'])
        blkWeight.grid(row=0, column=3, padx=3, pady=1)

        # Labelframe for Defects Data
        ####################################################################################################

        self.dataCont = LabelFrame(self.frame, bd=5, relief=FLAT)
        self.dataCont.grid(row=1, column=0, pady=1, padx=10)

        # Block No 
        blockTxt = Label(self.dataCont, font=self.font['S'], text="Block No")
        blockTxt.grid(row=0, column=0, pady=1, padx=15, sticky=W)

    ####################################################################################################
    """
    Functions Used for the Widgets Above
    """ 
    ####################################################################################################

    def load_SumData(self,root):
        if os.path.exists(self.filePath):
            wb = load_workbook(self.filePath)
            ws = wb.active

            maxcol = ws.max_column
            maxrow = ws.max_row

            xVal = 300 if maxcol*90 < 300 else maxcol*90
            if self.Wscreen < xVal*1.1: self.root.state('zoomed')
            else: self.root.geometry(str(xVal)+f"x{int(self.Hscreen*0.85)}+0+0")
            
            Separator(self.frame, orient="horizontal").grid(row=0, column=0, columnspan=maxcol, sticky=EW+S)

            # Store labels for highlighting
            self.defectLabels = {}  # {row_index: [list of labels in that row]}

            for i in range(maxcol):
                if i != 0:             
                    Button(self.dataCont, text="x", font=self.font['S'], height=1, width=2, bg="#fa6464", command= lambda i=i: self.delete(ws, wb, i,root)).grid(row=maxrow+1,column=i,pady=5,padx=15,sticky=S)
                    Label(self.dataCont, font=self.font['S'], text="Block "+str(i)).grid(row=0,column=i,pady=1,padx=15,sticky=E)
                else:
                    Button(self.dataCont, text="Confirm", font=self.font['S'], width=10, bg="#3085d6", command=lambda: self.createPRASS()).grid(row=maxrow+1,column=i,pady=5,padx=15,sticky=W)

                for j in range(maxrow):
                    exText = str(ws.cell(row=j+1,column=i+1).value)
                    label = Label(self.dataCont, font=self.font['S'], text=exText)
                    label.grid(row=j+1,column=i, pady=1, padx=15, sticky=W)
                    
                    # Store label reference for highlighting
                    if j not in self.defectLabels:
                        self.defectLabels[j] = []
                    self.defectLabels[j].append(label)
            
            # Check defect ratios after loading all data
            self.checkDefectRatios(ws, maxcol, maxrow)
    
    def checkDefectRatios(self, ws, maxcol, maxrow):
        """Check defect ratio for each defect mode and highlight if > 5%"""
        try:
            inqty = float(self.inData[3])  # inData[3] is inQty
            if inqty == 0:
                return  # Avoid division by zero
            
            outOfSpecDefects = []
            
            # Calculate total defects for each defect mode (row)
            for j in range(maxrow):
                defectMode = str(ws.cell(row=j+1, column=1).value)  # Defect mode name in column 1
                
                # Skip if it's not a valid defect mode (e.g., empty or None)
                if defectMode == "None" or defectMode == "":
                    continue
                
                # Sum defects across all blocks (columns 2 onwards in Excel, which is column index 2+)
                totalDefects = 0
                for col in range(2, maxcol + 1):  # Excel columns are 1-indexed, start from column 2
                    cellValue = ws.cell(row=j+1, column=col).value
                    if cellValue is not None:
                        try:
                            totalDefects += float(cellValue)
                        except (ValueError, TypeError):
                            pass
                
                # Calculate defect ratio
                if totalDefects > 0:
                    defectRatio = (totalDefects / inqty) * 100  # Convert to percentage
                    
                    # If ratio > 5%, highlight the row
                    if defectRatio > 5.0:
                        outOfSpecDefects.append((defectMode, defectRatio, totalDefects))
                        # Highlight all labels in this row
                        if j in self.defectLabels:
                            for label in self.defectLabels[j]:
                                label.config(bg="#ffcccc")  # Light red background
            
            # Show messagebox if any defect mode is out of spec
            if outOfSpecDefects:
                defectList = "\n".join([f"- {mode}: {ratio:.2f}% ({int(total)} defects)" 
                                       for mode, ratio, total in outOfSpecDefects])
                messagebox.showwarning(
                    "Out of Spec alert : Please Issue NCR",
                    f"The following defect modes exceed 5%:\n\n{defectList}\n\nPlease Issue NCR before continuing.",
                    parent=self.root
                )
                
                # Show popup prompt to send email alert
                send_email = messagebox.askyesno(
                    "Email Alert",
                    f"Defect ratios exceed 5% threshold.\n\nWould you like to send an email alert before proceeding?",
                    parent=self.root
                )
                
                if send_email:
                    self.sendDefectRatioEmailAlert(outOfSpecDefects)
        except (ValueError, TypeError, IndexError) as e:
            # If there's an error calculating ratios, just continue
            print(f"Error checking defect ratios: {e}")
    
    def sendDefectRatioEmailAlert(self, outOfSpecDefects):
        """Send email alert for defect ratios exceeding 5%"""
        try:
            # Load email configuration from settings.json
            if not self.email:
                messagebox.showerror(
                    "Email Configuration Error",
                    "Email configuration not found in settings.json. Please configure email settings.",
                    parent=self.root
                )
                return
            
            # Create email configuration DataFrame from settings
            email_data = [
                {
                    "RTO0006": self.email.get("RTO0006", "BAC"),
                    "RTO0010": self.email.get("RTO0010", "sendemail"),
                    "RTO0013_01": self.email.get("RTO0013_01", ""),  # Sender
                    "RTO0013_02": self.email.get("RTO0013_02", ""),  # Receiver
                    "RTO0013_03": self.email.get("RTO0013_03", ""),  # CC
                }
            ]
            emaildf = pd.DataFrame(email_data)
            
            # Validate email configuration
            if not email_data[0]["RTO0013_01"] or not email_data[0]["RTO0013_02"]:
                messagebox.showerror(
                    "Email Configuration Error",
                    "Email sender or receiver not configured in settings.json. Please update email settings.",
                    parent=self.root
                )
                return
            
            application_id = email_data[0]["RTO0006"]
            
            # Prepare defect details for email
            defect_details = {
                "Lot Number": self.inData[0],
                "Machine Number": self.inData[1],
                "Input Quantity": self.inData[3],
                "Block Weight": self.inData[4],
            }
            
            # Add defect information
            defect_info = []
            for mode, ratio, total in outOfSpecDefects:
                defect_info.append(f"{mode}: {ratio:.2f}% ({int(total)} defects)")
            defect_details["Defect Modes Exceeding 5%"] = "\n".join(defect_info)
            
            # Send email alert
            error_type = f"NCR Alert - Defect Ratio Exceeds 5% (Lot: {self.inData[0]})"
            sendErrorAlertEmail(
                emaildf,
                "sendemail",
                error_type,
                application_id,
                defect_details
            )
            
            messagebox.showinfo(
                "Email Sent",
                "Email alert has been sent successfully.",
                parent=self.root
            )
        except Exception as e:
            messagebox.showerror(
                "Email Error",
                f"Failed to send email alert: {str(e)}",
                parent=self.root
            )
            print(f"Error sending email alert: {e}")
        
    def delete(self,ws,wb,i,root):
        ws.delete_cols(i+1,1)
        wb.save(self.filePath)
        self.root.destroy()
        self.reset(root)

    def createPRASS(self):
        self.res = PRASS(self.root,self.inData,self.filePath,self.Wscreen,self.Hscreen).res
        self.root.destroy()
        self.root.quit()